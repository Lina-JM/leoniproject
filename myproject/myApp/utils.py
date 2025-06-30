import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from django.conf import settings


# Styling constants
RED_FILL = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')
GRAY_FILL = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
HEADER_FILL = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
BOLD_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal="center")
THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
THICK_BORDER = Border(left=Side(style='medium'),
                      right=Side(style='medium'),
                      top=Side(style='medium'),
                      bottom=Side(style='medium'))


def detect_header_row(file_path, expected_headers, max_rows=20):
    preview = pd.read_excel(file_path, header=None, nrows=max_rows)
    for i, row in preview.iterrows():
        row_values = row.astype(str).str.lower().tolist()
        match_count = sum(
            any(expected.lower() in cell for cell in row_values) for expected in expected_headers
        )
        if match_count >= 5:
            return i
    raise ValueError("Could not find header row containing expected headers.")


def read_excel_with_detected_header(file_path):
    expected_headers = ["Path", "Part number", "Plant", "Customer Part No", "Harness Description",
                       "Supplier No.", "Component No.", "Wire Number", "Mat. Group",
                       "Description", "UOM", "Req. Qty"]
    header_row = detect_header_row(file_path, expected_headers)
    df = pd.read_excel(file_path, header=header_row)
    df = df.dropna(how='all')
    return df


def clean_value(val):
    return str(val).strip().upper().replace('\xa0', '')


def build_data_dict(df, component_col=6, customer_col=3, quantity_col=11, description_col=9):
    return {
        (clean_value(row[component_col]), clean_value(row[customer_col])): {
            'Quantity': str(row[quantity_col]).strip(),
            'Description': str(row[description_col]).strip()
        }
        for _, row in df.iterrows()
        if pd.notna(row[component_col]) and pd.notna(row[customer_col])
    }


def apply_header_styles(ws):
    for row_num in [1, 2]:
        for col in range(1, 10):
            cell = ws.cell(row=row_num, column=col)
            cell.font = BOLD_FONT
            cell.alignment = CENTER_ALIGN
            cell.fill = GRAY_FILL if col == 5 else HEADER_FILL
            cell.border = THICK_BORDER


def set_column_widths(ws):
    column_widths = {
        1: 20,  # Component (X)
        2: 20,  # Customer Part (X)
        3: 15,  # Quantity (X)
        4: 40,  # Description (X)
        5: 5,   # Separator
        6: 20,  # Component (X-N)
        7: 20,  # Customer Part (X-N)
        8: 15,  # Quantity (X-N)
        9: 40   # Description (X-N)
    }
    for col, width in column_widths.items():
        ws.column_dimensions[chr(64 + col)].width = width


def generate_output(file1_path, file2_path, file):
    file.refresh_from_db()  # Reload the latest values from the database

    df1 = read_excel_with_detected_header(file1_path)
    df2 = read_excel_with_detected_header(file2_path)

    if df1.columns.tolist() != df2.columns.tolist():
        raise ValueError("Headers in BOOM(X) and BOOM(X-N) do not match.")

    data1 = build_data_dict(df1)
    data2 = build_data_dict(df2)

    all_keys = set(data1.keys()).union(data2.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    # Get just the week numbers (KW)
    kw1 = file.date1.isocalendar()[1]
    kw2 = file.date2.isocalendar()[1]

    # Header Row 1 - Only show KW numbers without dates
    ws.cell(row=1, column=1, value=f"KW {kw1}").fill = HEADER_FILL
    ws.cell(row=1, column=1).font = BOLD_FONT
    ws.cell(row=1, column=1).alignment = CENTER_ALIGN
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    ws.cell(row=1, column=5, value="").fill = GRAY_FILL  # Separator column

    ws.cell(row=1, column=6, value=f"KW {kw2}").fill = HEADER_FILL
    ws.cell(row=1, column=6).font = BOLD_FONT
    ws.cell(row=1, column=6).alignment = CENTER_ALIGN
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=9)

    # Header Row 2
    headers = ["Component (X)", "Customer Part (X)", "Quantity (X)", "Description (X)", "",
               "Component (X-N)", "Customer Part (X-N)", "Quantity (X-N)", "Description (X-N)"]
    ws.append(headers)

    apply_header_styles(ws)
    set_column_widths(ws)

    # Compare data and fill rows
    row_index = 3
    for key in sorted(all_keys):
        c, cp = key
        q1 = data1.get(key, {}).get('Quantity', '')
        d1 = data1.get(key, {}).get('Description', '')
        q2 = data2.get(key, {}).get('Quantity', '')
        d2 = data2.get(key, {}).get('Description', '')

        row = [
            c if key in data1 else "",
            cp if key in data1 else "",
            q1 if key in data1 else "",
            d1 if key in data1 else "",
            "",
            c if key in data2 else "",
            cp if key in data2 else "",
            q2 if key in data2 else "",
            d2 if key in data2 else ""
        ]

        ws.append(row)
        ws.cell(row=row_index, column=5).fill = GRAY_FILL  # Separator

        if key in data1 and key in data2:
            if q1 != q2:
                for col in [1, 2, 3, 6, 7, 8]:  # Component, Customer Part, Quantity for X and X-N
                    ws.cell(row=row_index, column=col).fill = RED_FILL

        elif key in data1 and key not in data2:
            if c:
                ws.cell(row=row_index, column=1).fill = ORANGE_FILL  # Component (X)
                ws.cell(row=row_index, column=4).fill = ORANGE_FILL  # Description (X)

        elif key in data2 and key not in data1:
            if c:
                ws.cell(row=row_index, column=6).fill = ORANGE_FILL  # Component (X-N)
                ws.cell(row=row_index, column=9).fill = ORANGE_FILL  # Description (X-N)

        row_index += 1

    # Add thin borders to all data cells
    for row in ws.iter_rows(min_row=3):  # Start from data rows
        for cell in row:
            cell.border = THIN_BORDER

    # Save
    output_filename = f"Comparison_Sitz_Rechts_VE_from_KW{kw1}_to_KW{kw2}.xlsx"
    output_filename = re.sub(r'[^\w\s\-_\[\]]', '', output_filename).strip()

    output_dir = os.path.join(settings.MEDIA_ROOT, 'outputs')
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, output_filename)

    wb.save(output_path)
    return output_path, output_filename
